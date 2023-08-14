from flask import Flask, request, send_file, send_from_directory
from flask_cors import CORS
import openpyxl, sqlite3, uuid, bcrypt, os

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

@app.route('/update', methods=['POST'])
def update_form():
    students = os.listdir('students')
    f = -1 
    for i, st in enumerate(students):
        if st.startswith(request.json['national_code']):
            f = i
            break

    if f != -1:
        return 'Found', 400

    # Load the workbook
    wb = openpyxl.load_workbook('form2.xlsx')

    # Get the worksheet
    ws = wb['form']
 
    # Name
    ws['B7'] = request.json['name']
    # LastName
    ws['I7'] = request.json['last_name'] 
    # Father Name
    ws['M7'] = request.json['father_name']
    # Field of Study
    ws['S7'] = request.json['field_of_study']
    # Number of Identity
    ws['X7'] = request.json['id_number']
    # (Alphabet) Number of Identity
    ws['AA7'] = request.json['id_alphabet']
    # (Int) Number of Identity
    ws['AA8'] = request.json['id_number_int']
    # Code Melli
    ws['I8'] = request.json['national_code']
    # Place of Issue
    ws['M8'] = request.json['issue_place']
    # Mother Name
    ws['I20'] = request.json['mother_name']
    # (Day) Birth
    ws['D9'] = request.json['birth_day']
    # (Month) Birth
    ws['F9'] = request.json['birth_month']
    # (Year) Birth
    ws['H9'] = request.json['birth_year']
    # Province
    ws['L9'] = request.json['province']
    # Country
    ws['P9'] = request.json['country']
    # City/Village
    ws['U9'] = request.json['city']
    # Postal Code
    ws['X9'] = request.json['postal_code']
    # Din
    ws['B10'] = request.json['din']
    # Religion
    ws['F10'] = request.json['religion']
    # Nationality
    ws['J10'] = request.json['nationality']
    # Status of body
    ws['N10'] = request.json['body_status'] 
    # Mobile number in "Shad" system
    ws['R10'] = request.json['shad_mobile']
    # Is Left-Handed?
    ws['Z11'] = request.json['left_handed']
    # Father's education
    ws['T18'] = request.json['father_education']
    # Father's occupation
    ws['X18'] = request.json['father_occupation']
    # Father's work address
    ws['L19'] = request.json['father_work_address']
    # Father's work phone
    ws['X19'] = request.json['father_work_phone']
    # (Day) Father's Birth
    ws['F19'] = request.json['father_birth_day']
    # (Month) Father's Birth  
    ws['H19'] = request.json['father_birth_month']
    # (Year) Father's Birth
    ws['J19'] = request.json['father_birth_year']
    # Father's ID
    ws['Q18'] = request.json['father_id']
    # Father's Place of Issue
    ws['Q19'] = request.json['father_issue_place'] 
    # Father's Insurance code
    ws['T19'] = request.json['father_insurance_code']
    # Father's National Code
    ws['M18'] = request.json['father_national_code']
    # Mother's education
    ws['T20'] = request.json['mother_education']
    # Mother's occupation
    ws['X20'] = request.json['mother_occupation']
    # Mother's work address
    ws['L21'] = request.json['mother_work_address']
    # Mother's work phone
    ws['X21'] = request.json['mother_work_phone']
    # (Day) Mother's Birth
    ws['F21'] = request.json['mother_birth_day']
    # (Month) Mother's Birth
    ws['H21'] = request.json['mother_birth_month']
    # (Year) Mother's Birth
    ws['J21'] = request.json['mother_birth_year']  

    # (Day) Supervisor's Birth
    ws['F21'] = request.json['supervisor_birth_day']
    # (Month) Supervisor's Birth
    ws['H21'] = request.json['supervisor_birth_month']
    # (Year) Supervisor's Birth
    ws['J21'] = request.json['supervisor_birth_year']

    # Supervisor name
    ws['I22'] = request.json['supervisor_name']
    # supervisor's education
    ws['T22'] = request.json['supervisor_education']
    # supervisor's occupation
    ws['X22'] = request.json['supervisor_occupation']
    # supervisor's work address
    ws['L23'] = request.json['supervisor_work_address']
    # supervisor's work phone
    ws['X19'] = request.json['supervisor_work_phone']
    # (Day) supervisor's Birth
    ws['F23'] = request.json['supervisor_birth_day']
    # (Month) supervisor's Birth  
    ws['h23'] = request.json['supervisor_birth_month']
    # (Year) supervisor's Birth
    ws['J23'] = request.json['supervisor_birth_year']
    # supervisor's ID
    ws['Q22'] = request.json['supervisor_id']
    # supervisor's Place of Issue
    ws['Q23'] = request.json['supervisor_issue_place'] 
    # supervisor's Insurance code
    ws['T23'] = request.json['supervisor_insurance_code']
    # supervisor's National Code
    ws['M22'] = request.json['supervisor_national_code']


    # Mother's ID
    ws['Q20'] = request.json['mother_id']
    # Mother's Place of Issue
    ws['Q21'] = request.json['mother_issue_place']
    # Mother's Insurance code
    ws['T21'] = request.json['mother_insurance_code']
    # Mother's National Code
    ws['M20'] = request.json['mother_national_code']
    # Address
    ws['H12'] = request.json['address']
    # Home Telephone
    ws['O12'] = request.json['home_phone']
    # Status of housing
    ws['Y12'] = request.json['housing_status']
    # Father's phone number
    # ws['E17'] = request.json['father_phone']
    # # Mother's phone number
    # ws['L17'] = request.json['mother_phone']  
    # Phone number
    ws['W10'] = request.json['phone']
    # Emergency phone number
    ws['S12'] = request.json['emergency_phone']
    # Who does the student live with at home?
    ws['J14'] = request.json['live_with']
    # The student's housing situation if he lives away from his family to study: 
    ws['T14'] = request.json['housing_situation']
    # Does the student have an independent study room:
    ws['AA14'] = request.json['study_room']
    # the number of family members:
    ws['E15'] = request.json['family_members']
    # How many children are there before him?
    ws['K15'] = request.json['children_before']
    # Who is the student supervisor?
    ws['Q15'] = request.json['student_supervisor']
    # Email
    ws['V15'] = request.json['email']
    # Height  
    ws['D11'] = request.json['height']
    # Weight
    ws['K11'] = request.json['weight']
    # Ability, skill and position or rank:
    ws['P11'] = request.json['ability']
    # The number used in the government portal
    ws['S8'] = request.json['gov_portal_number']
    # Status of pervious year
    ws['H23'] = request.json['previous_year_status']
    # The total GPA of the ninth grade: (Total GPA)
    ws['O16'] = request.json['ninth_gpa']
    # Accepted in "Nemoone-Dolati" Test
    ws['T16'] = request.json['dolati_test']
    # Pervious school name
    ws['I16'] = request.json['previous_school_name']  
    # Witness quota, sacrifice:
    ws['V13'] = request.json['witness_quota']
    # Under the cover of the Imam Khomeini (RA) relief committee:
    ws['J13'] = request.json['imam_relief']
    # Under welfare:
    ws['M13'] = request.json['welfare']
    # Does father work in Ministry of Education?
    ws['P13'] = request.json['father_education_ministry']
    # Does mother work in Ministry of Education?
    ws['S13'] = request.json['mother_education_ministry']

    # Father full name
    ws['I18'] = request.json['father_name'] + ' ' + request.json['last_name']

    # Save workbook
    wb.save('students/' + request.json['national_code'] + '-' + request.json['name'] + '-' + request.json['last_name'] + '-' + request.json['field_of_study'] + '.xlsx')

    return "Ok"



def get_db():
    db = sqlite3.connect('database.db')
    db.row_factory = sqlite3.Row


    return db

@app.route('/login', methods=['POST'])
def login():
    username = request.json['username']
    password = request.json['password']

    db = get_db()
    user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()

    if user and bcrypt.checkpw(bytes(password, 'utf-8'), user['password']):
        token = str(uuid.uuid4())
        db.execute('INSERT INTO tokens (user_id, token) VALUES (?, ?)',
                    (user['id'], token))
        db.commit()
        return {'token': token}
    else:
        return 'Invalid credentials', 401
        
@app.route('/students')  
def api():
    token = request.headers.get('X-Token') 
    if not token:
        return 'Unauthorized', 401
        
    db = get_db()
    result = db.execute('SELECT user_id FROM tokens WHERE token = ?', 
                        (token,)).fetchone()
                        
    if result:
        return os.listdir('students/')
    else:
        return 'Invalid token', 401

@app.route('/api')  
def test_api():
    token = request.headers.get('X-Token') 
    if not token:
        return 'Unauthorized', 401
        
    db = get_db()
    result = db.execute('SELECT user_id FROM tokens WHERE token = ?', 
                        (token,)).fetchone()
                        
    if result:
        return 'OK', 200
    else:
        return 'Invalid token', 401

@app.route('/student')  
def student():
    token = request.args.get('xToken') 
    if not token:
        return 'Unauthorized', 401
    db = get_db()
    result = db.execute('SELECT user_id FROM tokens WHERE token = ?', 
                        (token,)).fetchone()
    if result == None:
        return 'Unauthorized', 401
    students = os.listdir('students/')
    nCodes = []
    for st in students:
        nCodes.append(st.split('-')[0])
    user_national_code = request.args.get('nCode')
                        
    if user_national_code in nCodes :
        f = -1 
        for i, st in enumerate(students):
            if st.startswith(user_national_code):
                f = i
                break

        if f == -1:
            return 'Not Found', 404
        
        # return send_file(f'students/{os.listdir("students/")[f]}')
        return send_from_directory('students/', os.listdir("students/")[f], as_attachment=True, )
    else:
        return 'Not Found', 404
        
if __name__ == '__main__':
    app.run()